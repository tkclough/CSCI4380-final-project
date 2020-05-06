import psycopg2
import psycopg2.extras
import tarfile
import io
import re
import tqdm
import xlrd
import openpyxl
import gzip
import csv
import itertools
import math
import time
connection_string = "host='localhost' dbname='dbms_final_project' user='dbms_project_user' password='dbms_password'"


def main():
    start_time = time.time()

    with psycopg2.connect(connection_string) as conn:
        print("Writing schema")
        with open("schema.sql", "r") as f, conn.cursor() as cur:
            cur.execute(f.read())
        conn.commit()

        print("Loading data")

        # write to a new csv that only includes rows from 2008
        print('Loading measurement data')
        load_measurements(conn, 'datasets/HPD_v02r02_POR_s19400101_e20200422_c20200429.tar.gz')

        # read incident data
        print('Loading 2008 fire incident data')
        load_incident_address(conn, 'datasets/2008incidentaddress.xlsx')

        # compute closest stations
        print('Computing closest stations')
        compute_closest_stations(conn)

        print('Loading basic incident data')
        load_basic_incident(conn, 'datasets/basicincident.xlsx')

    print('Done in {:6.2f}s'.format(time.time() - start_time))


def batch_insert(conn, stmt, records, batch_size, page_size):
    """Insert a large number of values into a table in batches, with progress bar included."""
    with conn.cursor() as cur:
        record_iter = iter(records)
        for i in tqdm.tqdm(range(0, len(records), batch_size)):
            batch = itertools.islice(record_iter, min(batch_size, len(records) - i))
            psycopg2.extras.execute_values(cur, stmt, batch, page_size=page_size)

    conn.commit()


def load_basic_incident(conn, filename, batch_size=10000):
    """The basic incident file contains information about the events, not including the address info. For whatever
    reason, someone in the government thought xlsx files would be good for storing a lot of information, so it takes a
    while."""
    header = [
        'STATE', 'FDID', 'INC_DATE', 'INC_NO', 'EXP_NO', 'VERSION', 'DEPT_STA', 'INC_TYPE', 'ADD_WILD', 'AID', 'ALARM',
        'ARRIVAL', 'INC_CONT', 'LU_CLEAR', 'SHIFT', 'ALARMS', 'DISTRICT', 'ACT_TAK1', 'ACT_TAK2', 'ACT_TAK3', 'APP_MOD',
        'SUP_APP', 'EMS_APP', 'OTH_APP', 'SUP_PER', 'EMS_PER', 'OTH_PER', 'RESOU_AID', 'PROP_LOSS', 'CONT_LOSS',
        'PROP_VAL', 'CONT_VAL', 'FF_DEATH', 'OTH_DEATH', 'FF_INJ', 'OTH_INJ', 'DET_ALERT', 'HAZ_REL', 'MIXED_USE',
        'PROP_USE', 'CENSUS'
    ]
    nameToCol = {name: i for i, name in enumerate(header)}
    indices = [nameToCol[name] for name in ('INC_DATE', 'INC_NO', 'PROP_LOSS')]

    print('Opening excel workbook, give me a minute...')
    # for some reason doesn't define __enter__ so can't use context manager
    # using openpyxl here because it allows for constant memory access, even to large files like this one
    wb = openpyxl.load_workbook(filename=filename, read_only=True)
    try:
        with conn.cursor() as cur:
            ws = wb['basicincident']
            rows = iter(ws.rows)
            next(rows)
            for _ in tqdm.tqdm(range(0, ws.max_row - 1, batch_size), total=ws.max_row // batch_size):
                batch = itertools.islice(rows, batch_size)
                records = []
                for row in batch:
                    record = [(row[i].value if row[i].value != 'None' else None) for i in indices]

                    d = record[0]
                    if d:
                        d = int(d)
                        year = d % 10000
                        month = d // 1000000
                        day = (d // 10000) % 100
                        record[0] = '{:4d}-{:02d}-{:02d}'.format(year, month, day)

                    records.append(record)
                psycopg2.extras.execute_values(cur, "INSERT INTO basic_incident VALUES %s;", records, page_size=100)
                conn.commit()
    finally:
        wb.close()


def load_incident_address(conn, filename):
    """The incident address file stores location information about events, which can be referenced from the basic
    incident table."""
    header = [
        'STATE', 'FDID', 'INC_DATE', 'INC_NO', 'EXP_NO', 'LOC_TYPE', 'NUM_MILE', 'STREET_PRE', 'STREETNAME',
        'STREETTYPE', 'STREETSUF', 'APT_NO', 'CITY', 'STATE_ID', 'ZIP5', 'ZIP4', 'X_STREET'
    ]
    nameToCol = {name: i for (i, name) in enumerate(header)}

    print('Opening excel workbook, give me a minute...')
    records = []

    # xlrd is faster than openpyxl when the data fits in memory, like this file
    with xlrd.open_workbook(filename) as f:
        sheet = f.sheets()[0] # for some reason get_sheet(0) fails
        rows = sheet.get_rows()

        # skip header
        next(rows)
        for row in tqdm.tqdm(rows, total=sheet.nrows):
            data = [row[nameToCol[name]].value for name in ('STATE', 'FDID', 'INC_NO', 'CITY', 'ZIP5')]
            data = [(x if x != '' else None) for x in data]

            records.append(data)

    print('Inserting records...')
    batch_insert(conn, "INSERT INTO incident_address VALUES %s;", records, 10000, 100)


def load_measurements(conn, tarfilename, year=2008):
    """This loads precipitation data from a tar archive. It takes a while."""
    pat = r'(\d{4})-(\d{2})-(\d{2})'

    headers = [
        'StnID', 'Lat', 'Lon', 'Elev', 'Year-Month-Day', 'Element', 'HR00Val', 'HR00MF', 'HR00QF', 'HR00S1', 'HR00S2',
        'HR01Val', 'HR01MF', 'HR01QF', 'HR01S1', 'HR01S2', 'HR02Val', 'HR02MF', 'HR02QF', 'HR02S1', 'HR02S2', 'HR03Val',
        'HR03MF', 'HR03QF', 'HR03S1', 'HR03S2', 'HR04Val', 'HR04MF', 'HR04QF', 'HR04S1', 'HR04S2', 'HR05Val',
        'HR05MF', 'HR05QF', 'HR05S1', 'HR05S2', 'HR06Val', 'HR06MF', 'HR06QF', 'HR06S1', 'HR06S2', 'HR07Val', 'HR07MF',
        'HR07QF', 'HR07S1', 'HR07S2', 'HR08Val', 'HR08MF', 'HR08QF', 'HR08S1', 'HR08S2', 'HR09Val', 'HR09MF', 'HR09QF',
        'HR09S1', 'HR09S2', 'HR10Val', 'HR10MF', 'HR10QF', 'HR10S1', 'HR10S2', 'HR11Val', 'HR11MF', 'HR11QF', 'HR11S1',
        'HR11S2', 'HR12Val', 'HR12MF', 'HR12QF', 'HR12S1', 'HR12S2', 'HR13Val', 'HR13MF', 'HR13QF', 'HR13S1', 'HR13S2',
        'HR14Val', 'HR14MF', 'HR14QF', 'HR14S1', 'HR14S2', 'HR15Val', 'HR15MF', 'HR15QF', 'HR15S1', 'HR15S2', 'HR16Val',
        'HR16MF', 'HR16QF', 'HR16S1', 'HR16S2', 'HR17Val', 'HR17MF', 'HR17QF', 'HR17S1', 'HR17S2', 'HR18Val', 'HR18MF',
        'HR18QF', 'HR18S1', 'HR18S2', 'HR19Val', 'HR19MF', 'HR19QF', 'HR19S1', 'HR19S2', 'HR20Val', 'HR20MF', 'HR20QF',
        'HR20S1', 'HR20S2', 'HR21Val', 'HR21MF', 'HR21QF', 'HR21S1', 'HR21S2', 'HR22Val', 'HR22MF', 'HR22QF', 'HR22S1',
        'HR22S2', 'HR23Val', 'HR23MF', 'HR23QF', 'HR23S1', 'HR23S2', 'DlySum', 'DlySumMF', 'DlySumQF', 'DlySumS1',
        'DlySumS2'
    ]
    nameToCol = {name: i for i, name in enumerate(headers)}

    # open the output file for writing and the .tar.gz archive for reading
    print('Opening tarfile, give me a minute...')
    records = []
    skipped = []
    with tarfile.open(tarfilename) as t:
        # for each csv file in the archive
        for member in tqdm.tqdm(t.getmembers()):
            # open the csv file within the archive (binary), and wrap in line based layer
            found = False
            with t.extractfile(member) as f2, io.TextIOWrapper(f2, encoding='utf-8', newline='') as csvfile:
                next(csvfile)
                # for each line in the file
                for i, line in enumerate(csvfile):
                    parts = line.split(',')
                    if len(parts) != len(headers):
                        skipped.append((i, member.name))
                    # the date is the 4th column
                    date = parts[nameToCol['Year-Month-Day']].strip()
                    match = re.match(pat, date)

                    # determine if the year is correct, and if so, write
                    if match and int(match.group(1)) == year:
                        found = True
                        for hr in range(24):
                            colname = 'HR%02dVal' % (hr,)
                            val = parts[nameToCol[colname]]
                            dt = '%s %02d:00:00' % (date, hr)
                            records.append((parts[nameToCol['StnID']], dt, val))
                    elif found:
                        # once we're past the 2008 records, we won't find any more
                        break

    with conn.cursor() as cur:
        psycopg2.extras.execute_values(cur, "INSERT INTO measurement_skipped_rows VALUES %s;", skipped)

    print('Inserting records...')
    batch_insert(conn, "INSERT INTO measurement VALUES %s;", records, 100000, 1000)


def compute_closest_stations(conn):
    """In order to join the precipitation data and the fire data, we compute the closest stations to each zipcode."""
    records = []

    stations = list()
    with gzip.open('datasets/HPD_v02r02_stationinv_c20200429.csv.gz', 'rt') as station_file:
        csv_reader = csv.reader(station_file, delimiter=',')
        next(csv_reader)
        for row in csv_reader:
            stations.append((str(row[0]), float(row[1]), float(row[2])))
    stations.sort(key=lambda x: x[1])
    with open('datasets/raw', 'r') as zipcode_file:
        csv_reader = csv.reader(zipcode_file, delimiter=',')
        next(csv_reader)
        for row in tqdm.tqdm(csv_reader):
            zipcode = str(row[0])
            lat = float(row[1])
            lng = float(row[2])
            closest = ""
            dist = -1
            for station in stations:
                st_name = station[0]
                st_lat = station[1]
                st_lng = station[2]
                if dist >= 0 and (st_lat - lat) > dist:
                    break
                st_dist = math.sqrt((st_lat-lat)**2 + (st_lng-lng)**2)
                if st_dist < dist or dist == -1:
                    closest = st_name
                    dist = st_dist
            records.append((zipcode, closest))

    print('Inserting records...')
    batch_insert(conn, "INSERT INTO closest_station VALUES %s;", records, 1000, 100)


if __name__ == "__main__":
    main()
